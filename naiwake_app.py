"""
勘定科目内訳書作成ツール
対象: 地代家賃 / 雑収入（医業外収益）
入力:
  [Excel] JDL元帳Excel（列位置で判定）
    0列目: 年月日 / 1列目: 相手科目 / 2列目: 摘要 / 3列目: 借方 / 4列目: 貸方 / 5列目: 残高
  [CSV] 会計CSV（Shift-JIS）
    1列目が " #" で始まる行のみ処理
    2列目: 日付 / 7列目: 摘要（全角スペース区切り） / 14列目: 借方 / 15列目: 貸方
"""

import csv
import io
import math
import re
import unicodedata
import difflib

import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# 定数
# ─────────────────────────────────────────────
# JDL元帳Excelの列インデックス（0始まり）
COL_DESC   = 2  # 摘要（支払先＋内容、スペース2つ以上区切り）
COL_DEBIT  = 3  # 借方
COL_CREDIT = 4  # 貸方

CSV_ROW_MARKER = " #"  # 処理対象行の1列目プレフィックス（会計CSV）

# 摘要列にこれらのキーワードを含む行は金額に関わらずスキップ
SKIP_KEYWORDS = ["消費税額振替", "月計", "決算月計", "繰越", "期首残高"]

# 年月パターン（例: 2024年7月分 / 10月分）→ parts[n] がこれに一致したら取引内容としてスキップ
YEARMONTH_RE = re.compile(r'^(?:\d{4}年)?\d{1,2}月分')

# 取引内容のクリーニング用正規表現
_CONTENT_YEARMONTH_FULL_RE  = re.compile(r'^(?:\d{4}年)?\d{1,2}月分\s*$')   # 全体が年月のみ
_CONTENT_YEARMONTH_STRIP_RE = re.compile(r'\d{4}年\d{1,2}月分')              # YYYY年M月分 を除去


def _clean_content(content: str) -> str:
    """取引内容から年月パターンを除去する。
    - 全体が「YYYY年M月分」または「M月分」→ 空文字を返す
    - 「YYYY年M月分」を含む場合 → その部分を除去して返す
    """
    if _CONTENT_YEARMONTH_FULL_RE.match(content):
        return ""
    return _CONTENT_YEARMONTH_STRIP_RE.sub("", content).strip()

# 前期計上分戻入 / 当期計上分 の判定キーワード（摘要の parts[0] に対して照合）
_MAE_MODOSHI_KWS = frozenset({"前期計上分戻入", "前期分戻入", "前期末未収金戻入"})
_TOUKI_KWS       = frozenset({"当期計上分", "当期分", "前期末未収金", "期末未収金", "期末計上分"})


def _is_mae_modoshi(p0_ns: str) -> bool:
    """parts[0]のスペース除去版が前期計上分戻入系かどうか判定する。"""
    return any(kw in p0_ns for kw in _MAE_MODOSHI_KWS) or p0_ns.startswith("前期分")


def _is_touki(p0_ns: str) -> bool:
    """parts[0]のスペース除去版が当期計上分系かどうか判定する（戻入系より後に評価すること）。"""
    return any(kw in p0_ns for kw in _TOUKI_KWS)

MODES = {
    "地代家賃": {
        "title": "地代家賃の内訳",
        "amount_col_idx": COL_DEBIT,
        "opposite_col_idx": COL_CREDIT,
        "csv_amount_col_name": "借方金額",
        "csv_opposite_col_name": "貸方金額",
        "key_label": "支払先（貸主）名",
        "amount_label": "支払金額",
        "default_top_n": 0,                    # 0 = 全件表示
        "taxable_kaku_codes": {"31", "32", "33"},  # 課区がこれらなら税込み
        "group_map": {                          # 支払先名の強制統合マップ
            "半場新一": "半場進一",              # 誤字修正
            "加藤":   "加藤純",                  # 名前の省略形を正式名称に統一
            "ゴキタジュンコ（羽鳥菜佑家賃）": "ゴキタジュンコ",  # 括弧付き注記を除去
        },
        "group_by_content": False,             # 支払先名のみで集計
        # 残高チェック基準値：
        #   税抜き法人 → 「合計金額」行の借方合計（消費税振替後の残高は使えないため）
        #   税込み法人 → 最終残高（元帳終了直前の残高列の値）
        "ledger_tolerance": 1_000,
    },
    "雑収入（医業外収益）": {
        "title": "雑収入（医業外収益）の内訳",
        "amount_col_idx": COL_CREDIT,
        "opposite_col_idx": COL_DEBIT,
        "csv_amount_col_name": "貸方金額",
        "csv_opposite_col_name": "借方金額",
        "key_label": "収入先名",
        "amount_label": "収入金額",
        "default_top_n": 15,                   # 上位15件＋その他
        "taxable_kaku_codes": {"11"},           # 課区がこれなら税込み
        "group_map": {                          # 収入先名の強制統合マップ
            "東洋リネンサプライ": "東洋リネン",
        },
        "group_by_content": True,              # 収入先名＋取引内容の組み合わせで集計
        # 残高チェック基準値：
        #   税抜き法人 → 「合計金額」行の貸方合計（消費税振替後の残高は使えないため）
        #   税込み法人 → 最終残高（元帳終了直前の残高列の値）
        "ledger_tolerance": 1_000,
    },
}

SIMILARITY_THRESHOLD = 0.80  # 表記ゆれとみなす類似度の閾値
CONTENT_COL = "取引内容"    # 摘要の2番目の部分（支払先ごとの代表内容）


# ─────────────────────────────────────────────
# 文字正規化
# ─────────────────────────────────────────────
def hankaku_to_zenkaku(text: str) -> str:
    """半角カナを全角カナに変換する（例：ﾚｵﾊﾟﾚｽ21→レオパレス21）"""
    if not text:
        return text

    # 半角カナ→全角カナ 基本変換テーブル
    HK_MAP = {
        'ｦ': 'ヲ', 'ｧ': 'ァ', 'ｨ': 'ィ', 'ｩ': 'ゥ', 'ｪ': 'ェ', 'ｫ': 'ォ',
        'ｬ': 'ャ', 'ｭ': 'ュ', 'ｮ': 'ョ', 'ｯ': 'ッ', 'ｰ': 'ー',
        'ｱ': 'ア', 'ｲ': 'イ', 'ｳ': 'ウ', 'ｴ': 'エ', 'ｵ': 'オ',
        'ｶ': 'カ', 'ｷ': 'キ', 'ｸ': 'ク', 'ｹ': 'ケ', 'ｺ': 'コ',
        'ｻ': 'サ', 'ｼ': 'シ', 'ｽ': 'ス', 'ｾ': 'セ', 'ｿ': 'ソ',
        'ﾀ': 'タ', 'ﾁ': 'チ', 'ﾂ': 'ツ', 'ﾃ': 'テ', 'ﾄ': 'ト',
        'ﾅ': 'ナ', 'ﾆ': 'ニ', 'ﾇ': 'ヌ', 'ﾈ': 'ネ', 'ﾉ': 'ノ',
        'ﾊ': 'ハ', 'ﾋ': 'ヒ', 'ﾌ': 'フ', 'ﾍ': 'ヘ', 'ﾎ': 'ホ',
        'ﾏ': 'マ', 'ﾐ': 'ミ', 'ﾑ': 'ム', 'ﾒ': 'メ', 'ﾓ': 'モ',
        'ﾔ': 'ヤ', 'ﾕ': 'ユ', 'ﾖ': 'ヨ',
        'ﾗ': 'ラ', 'ﾘ': 'リ', 'ﾙ': 'ル', 'ﾚ': 'レ', 'ﾛ': 'ロ',
        'ﾜ': 'ワ', 'ﾝ': 'ン',
        '｡': '。', '｢': '「', '｣': '」', '､': '、', '･': '・',
    }
    DAKUTEN    = 'ﾞ'  # U+FF9E
    HANDAKUTEN = 'ﾟ'  # U+FF9F

    # 濁点付き変換テーブル
    DAKUTEN_MAP = {
        'カ': 'ガ', 'キ': 'ギ', 'ク': 'グ', 'ケ': 'ゲ', 'コ': 'ゴ',
        'サ': 'ザ', 'シ': 'ジ', 'ス': 'ズ', 'セ': 'ゼ', 'ソ': 'ゾ',
        'タ': 'ダ', 'チ': 'ヂ', 'ツ': 'ヅ', 'テ': 'デ', 'ト': 'ド',
        'ハ': 'バ', 'ヒ': 'ビ', 'フ': 'ブ', 'ヘ': 'ベ', 'ホ': 'ボ',
        'ウ': 'ヴ',
    }
    # 半濁点付き変換テーブル
    HANDAKUTEN_MAP = {
        'ハ': 'パ', 'ヒ': 'ピ', 'フ': 'プ', 'ヘ': 'ペ', 'ホ': 'ポ',
    }

    result = []
    i = 0
    while i < len(text):
        ch = text[i]
        if ch in HK_MAP:
            zk = HK_MAP[ch]
            if i + 1 < len(text):
                next_ch = text[i + 1]
                if next_ch == DAKUTEN and zk in DAKUTEN_MAP:
                    result.append(DAKUTEN_MAP[zk])
                    i += 2
                    continue
                elif next_ch == HANDAKUTEN and zk in HANDAKUTEN_MAP:
                    result.append(HANDAKUTEN_MAP[zk])
                    i += 2
                    continue
            result.append(zk)
        else:
            result.append(ch)
        i += 1

    return ''.join(result)


def normalize_text(text):
    if not text:
        return text
    # NFC正規化：Unicodeレベルで分離した濁点・半濁点を結合（最も確実な方法）
    text = unicodedata.normalize('NFC', text)
    # 上記で拾えない残余パターンを個別置換で補完
    replacements = [
        ('ハ\u309a', 'パ'), ('ヒ\u309a', 'ピ'), ('フ\u309a', 'プ'),
        ('ヘ\u309a', 'ペ'), ('ホ\u309a', 'ポ'),
        ('カ\u3099', 'ガ'), ('キ\u3099', 'ギ'), ('ク\u3099', 'グ'),
        ('ケ\u3099', 'ゲ'), ('コ\u3099', 'ゴ'),
        ('ハ\u3099', 'バ'), ('ヒ\u3099', 'ビ'), ('フ\u3099', 'ブ'),
        ('ヘ\u3099', 'ベ'), ('ホ\u3099', 'ボ'),
        # スタンドアロン濁点（U+309B）への対処
        # ホ゛→ボ（NFC では結合されない独立濁点を明示変換）
        ('ホ\u309b', 'ボ'),
        # ボ゛→ボ（ボは既に濁点済み、余分な独立/結合濁点を除去）
        ('ボ\u309b', 'ボ'), ('ボ\u3099', 'ボ'),
    ]
    for src, dst in replacements:
        text = text.replace(src, dst)
    # 斎藤系異体字を統一
    for v in ['斉', '斎']:
        text = text.replace(v, '齋')
    # 渡辺系異体字を統一（渡邉・渡邊 → 渡辺）
    text = text.replace('渡邉', '渡辺').replace('渡邊', '渡辺')
    # 括弧付き注記を除去（例：齋藤直永（...）→ 齋藤直永）
    text = re.sub(r'（[^）]*）', '', text).strip()
    return text


def calc_tax_excluded(
    amount: float,
    kaku_ku: str,
    zei_ku: str,
    taxable_kaku_codes: set[str],
) -> float:
    """
    課区・税区に基づいて税抜き金額を計算する（端数は math.floor で切り捨て）。
    - 課区が taxable_kaku_codes に含まれる場合のみ税抜き計算を適用する
    - 税区 "10" / "71" → ÷1.1（10%）、税区 "9" / "70" → ÷1.08（8%）
    - 上記以外の課区（非課税など）はそのまま返す
    """
    if kaku_ku.strip() not in taxable_kaku_codes:
        return amount
    zei = zei_ku.strip()
    if zei in ("10", "71"):
        return math.floor(amount / 1.1)
    elif zei in ("9", "70"):
        return math.floor(amount / 1.08)
    return amount


# ─────────────────────────────────────────────
# ユーティリティ
# ─────────────────────────────────────────────
def find_similar_groups(names: list[str], threshold: float) -> dict[str, list[str]]:
    """
    difflib を使って名称の表記ゆれをグルーピングする。
    戻り値: {代表名: [類似名リスト]} （類似名が1件のみのグループは除外）
    """
    remaining = list(names)
    groups: dict[str, list[str]] = {}

    while remaining:
        base = remaining.pop(0)
        similar = [base]
        not_matched = []

        for candidate in remaining:
            ratio = difflib.SequenceMatcher(None, base, candidate).ratio()
            if ratio >= threshold:
                similar.append(candidate)
            else:
                not_matched.append(candidate)

        if len(similar) > 1:
            groups[base] = similar

        remaining = not_matched

    return groups


def load_jdl_excel(
    uploaded,
    amount_col_idx: int,
    opposite_col_idx: int,
    key_label: str,
    amount_label: str,
    taxable_kaku_codes: set[str],
    group_map: dict[str, str] | None = None,
    skiprows: int = 0,
) -> tuple[pd.DataFrame, pd.DataFrame, bool, float | None]:
    """
    JDL元帳Excelを列位置で読み込み、整形済みDataFrameと生データと税抜きフラグと最終残高を返す。
    - 全シートを読み込んで結合（複数月シート対応）
    - 各シートのヘッダー行（1列目が「年月日」の行）から課区・税区列を検出
    - 摘要列（COL_DESC）をスペース2つ以上で分割し、先頭部分を支払先/収入先名とする
    - 全シートに「消費税額振替」の行が存在する場合（税抜き法人）のみ税抜き計算を適用する
    - 「消費税額振替」の行がない場合（税込み法人）は金額をそのまま使用する
    - 正列・逆列の差額（純額）を金額とする（返金・前期計上分戻入を自動でマイナス計上）
    - 純額が0の行は除外
    """
    # 全シートを読み込む（sheet_name=None で dict[シート名→DataFrame] を返す）
    sheets = pd.read_excel(uploaded, header=None, skiprows=skiprows, dtype=str, sheet_name=None)
    raw = pd.concat(sheets.values(), ignore_index=True)

    if raw.shape[1] <= max(COL_DESC, amount_col_idx, opposite_col_idx):
        raise ValueError(
            f"列数が不足しています（{raw.shape[1]}列）。"
            f"最低でも {max(COL_DESC, amount_col_idx, opposite_col_idx) + 1} 列必要です。"
        )

    # 各シートの列ヘッダー行を検出：1列目が「年月日」を含む行
    header_mask = raw.iloc[:, 0].fillna("").astype(str).str.contains("年月日", na=False)

    # ヘッダー行から課区・税区・残高の列インデックスを動的に取得
    col_kaku_ku_idx: int | None = None
    col_zei_ku_idx: int | None = None
    col_balance_idx: int | None = None
    header_rows = raw[header_mask]
    if not header_rows.empty:
        hrow = header_rows.iloc[0]
        for i, cell in enumerate(hrow):
            c = str(cell).strip()
            if c == "課区":
                col_kaku_ku_idx = i
            elif c == "税区":
                col_zei_ku_idx = i
            elif c == "残高":
                col_balance_idx = i

    # ヘッダー行を除去
    raw = raw[~header_mask].reset_index(drop=True)

    # 残高列の最終値を取得（元帳終了直前の残高）
    ledger_final_balance: float | None = None
    if col_balance_idx is not None and col_balance_idx < raw.shape[1]:
        balance_series = pd.to_numeric(raw.iloc[:, col_balance_idx], errors="coerce")
        valid_balances = balance_series.dropna()
        if not valid_balances.empty:
            ledger_final_balance = float(valid_balances.iloc[-1])

    # 税抜き計算フラグの決定
    # 全シートの摘要列をスキャンし「消費税額振替」の行があれば税抜き法人
    _desc_nospace = (
        raw.iloc[:, COL_DESC].fillna("").astype(str)
        .str.replace(r"\s+", "", regex=True)
    )
    apply_tax_exclusion: bool = _desc_nospace.str.contains("消費税額振替", na=False).any()

    desc_col = raw.iloc[:, COL_DESC].fillna("").astype(str).str.strip()

    _group_map = group_map or {}

    def extract_payee_and_content(desc: str) -> tuple[str, str]:
        parts = re.split(r"\s{2,}", desc)
        # 「〇月分給与」など摘要に「給与」を含む行はすべて「寮費」に統合
        if "給与" in desc:
            return "寮費", "寮費"
        p0_ns = re.sub(r"\s+", "", parts[0]) if parts else ""
        # 前期計上分戻入・当期計上分系はparts[0]がキーワードなのでparts[1]を支払先とする
        payee_idx = 1 if (_is_mae_modoshi(p0_ns) or _is_touki(p0_ns)) else 0
        payee = normalize_text(parts[payee_idx].strip()) if payee_idx < len(parts) else ""
        # parts[payee_idx+1] が年月パターンなら1つ後ろを取引内容として採用
        content_idx = payee_idx + 1
        if content_idx < len(parts) and YEARMONTH_RE.match(parts[content_idx].strip()):
            content_idx += 1
        content = _clean_content(normalize_text(parts[content_idx].strip()) if content_idx < len(parts) else "")
        payee = _group_map.get(payee, payee)
        return payee, content

    extracted    = desc_col.apply(extract_payee_and_content)
    payee_col    = extracted.apply(lambda x: x[0])
    content_col  = extracted.apply(lambda x: x[1])

    # 課区・税区列が検出できた場合かつ税抜き法人の場合に税抜き計算を適用する（主列・逆列ともに）
    # 税込み法人（apply_tax_exclusion=False）は生金額をそのまま使用する
    raw_amounts = pd.to_numeric(raw.iloc[:, amount_col_idx], errors="coerce").fillna(0)
    raw_opp_amounts = pd.to_numeric(raw.iloc[:, opposite_col_idx], errors="coerce").fillna(0)
    if apply_tax_exclusion and col_kaku_ku_idx is not None and col_zei_ku_idx is not None:
        kaku_ku_col = raw.iloc[:, col_kaku_ku_idx].fillna("").astype(str).str.strip()
        zei_ku_col  = raw.iloc[:, col_zei_ku_idx].fillna("").astype(str).str.strip()
        amount_col_data = pd.Series(
            [calc_tax_excluded(a, k, z, taxable_kaku_codes)
             for a, k, z in zip(raw_amounts, kaku_ku_col, zei_ku_col)],
            index=raw_amounts.index,
        )
        opp_col_data = pd.Series(
            [calc_tax_excluded(a, k, z, taxable_kaku_codes)
             for a, k, z in zip(raw_opp_amounts, kaku_ku_col, zei_ku_col)],
            index=raw_opp_amounts.index,
        )
    else:
        amount_col_data = raw_amounts
        opp_col_data = raw_opp_amounts

    # 純額 = 主列 − 逆列（返金・前期計上分戻入などを自動でマイナス計上）
    net_amounts = amount_col_data - opp_col_data

    df = pd.DataFrame({
        key_label:    payee_col,
        CONTENT_COL:  content_col,
        amount_label: net_amounts,
        "_desc":      desc_col,
    })

    # スキップキーワードを含む行を除外（月計・繰越・消費税額振替・決算など）
    # スペース除去後の文字列でマッチ → 「繰 越」「次 頁 へ 繰 越」などにも対応
    desc_no_space = desc_col.str.replace(r"\s+", "", regex=True)
    skip_pattern = "|".join(re.escape(kw) for kw in SKIP_KEYWORDS)
    df = df[~desc_no_space.str.contains(skip_pattern, na=False)]

    # 不要行の除外（純額0・支払先が空/nan文字列）
    df = df[df[amount_label] != 0]
    df = df[df[key_label].notna()]
    df = df[~df[key_label].isin(["", "nan", "NaN"])]

    df = df.drop(columns=["_desc"]).reset_index(drop=True)

    return df, raw, apply_tax_exclusion, ledger_final_balance



def load_csv_file(
    uploaded,
    amount_col_name: str,
    opposite_col_name: str,
    key_label: str,
    amount_label: str,
    taxable_kaku_codes: set[str],
    group_map: dict[str, str] | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame, float | None, float | None, float, bool]:
    """
    会計CSVファイル（Shift-JIS）を読み込み、以下の6値を返す。
      df                   : 整形済みDataFrame
      raw                  : 生データDataFrame
      ledger_final_balance : 「元帳終了」直前の残高列最終値（雑収入モード向け）
      ledger_total_row     : 「合計金額」行の主列値（地代家賃モード向け）
      taxinc_net_total     : 集計行の税込み純額の合計（地代家賃モード残高チェック用）
      apply_tax_exclusion  : 税抜き計算を適用するか（消費税額振替の行が決算月に存在する場合 True）

    集計ロジック（1列目が CSV_ROW_MARKER の行を対象）：
    ・スキップ行（SKIP_KEYWORDS に該当）→ 除外
    ・前期計上分戻入 → 逆列金額をマイナスとして計上
    ・当期計上分     → 主列金額をプラスとして計上
    ・通常行         → 主列 − 逆列 の純額を計上
    ・apply_tax_exclusion=True（税抜き法人）の場合のみ課区・税区に基づく税抜き計算を適用する
    """
    content = uploaded.read()
    text = content.decode('shift_jis', errors='replace')

    # csv モジュールで1行ずつ読み込み、ヘッダー行・データ行を振り分ける
    all_rows: list[list[str]] = []
    filtered_rows: list[list[str]] = []
    col_desc_idx: int | None = None
    col_amount_idx: int | None = None
    col_opposite_idx: int | None = None
    col_balance_idx: int | None = None
    col_kaku_ku_idx: int | None = None
    col_zei_ku_idx: int | None = None
    ledger_final_balance: float | None = None   # 残高列の最終値（元帳終了直前）
    ledger_total_row: float | None = None       # 「合計金額」行の主列値
    last_balance: float | None = None

    reader = csv.reader(io.StringIO(text))
    for row in reader:
        all_rows.append(row)

        # ヘッダー行の検出：「借方金額」または「貸方金額」を含む行（部分一致）
        if col_desc_idx is None and any(
            amount_col_name in cell or opposite_col_name in cell for cell in row
        ):
            for i, cell in enumerate(row):
                c = cell.strip()
                if c == "摘要":
                    col_desc_idx = i
                elif amount_col_name in c and col_amount_idx is None:
                    col_amount_idx = i
                elif opposite_col_name in c and col_opposite_idx is None:
                    col_opposite_idx = i
                elif c == "残高":
                    col_balance_idx = i
                elif c == "課区":
                    col_kaku_ku_idx = i
                elif c == "税区":
                    col_zei_ku_idx = i
            continue  # ヘッダー行自体はデータとして追加しない

        # 「合計金額」行の検出：主列（借方/貸方）の合計値を取得
        if any("合計金額" in cell for cell in row):
            if col_amount_idx is not None and col_amount_idx < len(row):
                val = row[col_amount_idx].replace(",", "").strip()
                if val:
                    try:
                        ledger_total_row = float(val)
                    except ValueError:
                        pass
            continue

        # 「元帳終了」行の検出：直前までの残高最終値を確定
        if any("元帳終了" in cell for cell in row):
            if last_balance is not None:
                ledger_final_balance = last_balance
            continue

        # 残高列の最終値を追跡（元帳終了直前の値が正解値になる）
        if col_balance_idx is not None and col_balance_idx < len(row):
            val = row[col_balance_idx].replace(",", "").strip()
            if val:
                try:
                    last_balance = float(val)
                except ValueError:
                    pass

        # データ行：1列目が CSV_ROW_MARKER で始まる行のみ収集
        if row and row[0].startswith(CSV_ROW_MARKER):
            filtered_rows.append(row)

    # 生データ表示用 DataFrame（列数が異なる行が混在するため最大列数に合わせてパディング）
    max_cols = max((len(r) for r in all_rows), default=0)
    raw = pd.DataFrame(
        [r + [""] * (max_cols - len(r)) for r in all_rows],
        dtype=str,
    )

    # ── 税抜き計算フラグの決定 ──
    # 摘要列に「消費税額振替」が含まれる行があれば税抜き法人（Excel側の検出と同一ロジック）
    # ※ 日付による決算月判定は日付フォーマット依存で誤検出があるため使用しない
    apply_tax_exclusion: bool = False
    if col_desc_idx is not None:
        for _row in filtered_rows:
            _desc_cell = re.sub(r"\s+", "", hankaku_to_zenkaku(
                _row[col_desc_idx].strip() if col_desc_idx < len(_row) else ""
            ))
            if "消費税額振替" in _desc_cell:
                apply_tax_exclusion = True
                break

    if not filtered_rows:
        return pd.DataFrame(), raw, ledger_final_balance, ledger_total_row, 0.0, apply_tax_exclusion

    if col_desc_idx is None:
        raise ValueError(
            "ヘッダー行（「借方金額」または「貸方金額」を含む行）が見つかりませんでした。"
        )
    if col_amount_idx is None:
        raise ValueError(
            f"ヘッダー行に「{amount_col_name}」列が見つかりませんでした。"
        )

    # ── ヘルパー：列インデックス指定でセルを float に変換 ──
    def _cell_float(row: list[str], idx: int | None) -> float:
        if idx is None or idx >= len(row):
            return 0.0
        val = row[idx].replace(",", "").strip()
        try:
            return float(val) if val else 0.0
        except ValueError:
            return 0.0

    # ── 1行ずつ処理して records に蓄積 ──
    skip_pattern = "|".join(re.escape(kw) for kw in SKIP_KEYWORDS)
    _group_map = group_map or {}
    records: list[dict] = []
    taxinc_net_total: float = 0.0  # 税込み純額の合計（地代家賃モード残高チェック用）

    for row in filtered_rows:
        # 摘要（半角カナ→全角カナ変換済み、スペース除去版も用意）
        desc_raw = row[col_desc_idx] if col_desc_idx < len(row) else ""
        desc = hankaku_to_zenkaku(desc_raw.strip())
        desc_ns = re.sub(r"\s+", "", desc)  # スペース除去版（キーワードマッチ用）

        # スキップ判定
        if re.search(skip_pattern, desc_ns):
            continue

        # 全角スペース区切りで分割し、parts[0] で前期/当期キーワードを判定
        parts = desc.split('\u3000')
        p0_ns = re.sub(r"\s+", "", parts[0]) if parts else ""
        _mae = _is_mae_modoshi(p0_ns)
        _tou = _is_touki(p0_ns) and not _mae

        # 課区・税区
        kaku = row[col_kaku_ku_idx].strip() if col_kaku_ku_idx is not None and col_kaku_ku_idx < len(row) else ""
        zei  = row[col_zei_ku_idx].strip()  if col_zei_ku_idx  is not None and col_zei_ku_idx  < len(row) else ""

        # 税込みの生金額（残高チェック用）
        main_raw = _cell_float(row, col_amount_idx)
        opp_raw  = _cell_float(row, col_opposite_idx)

        # 集計用金額：税抜き法人のみ課区・税区に基づいて税抜き計算を適用する
        # 税込み法人（apply_tax_exclusion=False）は生金額をそのまま使用する
        # ※ 前期計上分戻入・当期計上分の処理ロジック自体は apply_tax_exclusion によらず同一
        if apply_tax_exclusion:
            main_amount = calc_tax_excluded(main_raw, kaku, zei, taxable_kaku_codes)
            opp_amount  = calc_tax_excluded(opp_raw,  kaku, zei, taxable_kaku_codes)
        else:
            main_amount = main_raw
            opp_amount  = opp_raw

        # 金額区分（税抜き純額 / 税込み純額 を並行計算）
        if _mae:
            net        = -opp_amount   # 逆列をマイナスで計上（前期計上分戻入系）
            taxinc_net = -opp_raw
        elif _tou:
            net        = main_amount   # 主列をプラスで計上（当期計上分系）
            taxinc_net = main_raw
        else:
            net        = main_amount - opp_amount  # 通常行：純額
            taxinc_net = main_raw    - opp_raw

        # ── net==0 の処理 ──
        # 両方の生金額が 0（本当に空の行）→ スキップ
        if main_raw == 0 and opp_raw == 0:
            continue
        # 貸方に金額があるのに net==0 になった場合
        # （税抜き丸め誤差 / 列検出ズレなど）→ 貸方全額をマイナスで強制計上
        if net == 0 and opp_raw > 0:
            net        = -opp_amount
            taxinc_net = -opp_raw
        # 上記以外で net==0 （借方と貸方が完全に相殺） → スキップ
        elif net == 0:
            continue

        taxinc_net_total += taxinc_net

        # 支払先・取引内容を摘要から抽出（全角スペース区切り）
        # parts はループ冒頭で既に計算済み
        if "給与" in desc:
            payee, content = "寮費", "寮費"
        else:
            # 前期/当期キーワード行は parts[0]=キーワードなので parts[1] を支払先とする
            payee_idx   = 1 if (_mae or _tou) else 0
            payee       = normalize_text(parts[payee_idx].strip()) if payee_idx < len(parts) else ""
            # parts[payee_idx+1] が年月パターンなら1つ後ろを取引内容として採用
            content_idx = payee_idx + 1
            if content_idx < len(parts) and YEARMONTH_RE.match(parts[content_idx].strip()):
                content_idx += 1
            content = _clean_content(normalize_text(parts[content_idx].strip()) if content_idx < len(parts) else "")
            payee   = _group_map.get(payee, payee)

        if not payee or payee in ("nan", "NaN"):
            continue

        records.append({key_label: payee, CONTENT_COL: content, amount_label: net})

    df = (
        pd.DataFrame(records)
        if records
        else pd.DataFrame(columns=[key_label, CONTENT_COL, amount_label])
    )
    df = df.reset_index(drop=True)

    return df, raw, ledger_final_balance, ledger_total_row, taxinc_net_total, apply_tax_exclusion


def auto_merge_by_frequency(
    names: list[str],
    counts: dict[str, int],
    threshold: float = 0.75,
) -> dict[str, str]:
    """
    類似度 threshold 以上のグループで、出現件数が最多の名称を正式名称として採用する。
    戻り値: {旧名称: 正式名称}（正式名称自身のエントリは含まない）
    """
    # 件数の多い順に処理することで、先に処理された名称が正式名称になる
    sorted_names = sorted(names, key=lambda n: counts.get(n, 0), reverse=True)
    name_map: dict[str, str] = {}
    assigned: set[str] = set()

    for base in sorted_names:
        if base in assigned:
            continue
        for candidate in sorted_names:
            if candidate == base or candidate in assigned:
                continue
            # 一方が他方の前方一致の場合はマージしない
            # 例：「千葉県」と「千葉県医務国保」は別物
            if base.startswith(candidate) or candidate.startswith(base):
                continue
            ratio = difflib.SequenceMatcher(None, base, candidate).ratio()
            if ratio >= threshold:
                name_map[candidate] = base  # base が件数最多 → 正式名称
                assigned.add(candidate)
        assigned.add(base)

    return name_map


def collapse_to_top_n(df: pd.DataFrame, key_col: str, amount_col: str, top_n: int) -> pd.DataFrame:
    """上位 top_n 件を残し、残りを「その他（N件）」1行にまとめる。top_n=0 は全件表示。"""
    if top_n <= 0 or len(df) <= top_n:
        return df
    top = df.iloc[:top_n].copy()
    other_count = len(df) - top_n
    other_sum = df.iloc[top_n:][amount_col].sum()
    # 全列を揃えるため df の列構成に合わせて other_row を生成
    other_data = {col: [""] for col in df.columns}
    other_data[key_col]    = [f"その他（{other_count}件）"]
    other_data[amount_col] = [other_sum]
    other_row = pd.DataFrame(other_data, index=[top_n + 1])
    return pd.concat([top, other_row])


def aggregate(
    df: pd.DataFrame,
    key_col: str,
    amount_col: str,
    name_map: dict[str, str],
    group_by_content: bool = False,
    merge_only_payees: set[str] | None = None,
) -> pd.DataFrame:
    """
    name_map に従って名称を統一してから金額を集計する。

    group_by_content=True の場合は (key_col, CONTENT_COL) の組み合わせで集計し、
    収入先ごとに取引内容が異なれば別行として出力する。
    ただし merge_only_payees に含まれる収入先は key_col のみで集計する
    （GROUP_MAP で統合済みの名称など、取引内容を問わずまとめたい収入先向け）。

    group_by_content=False（デフォルト）の場合は key_col のみで集計し、
    CONTENT_COL が存在する場合はグループ内の最頻値を代表内容として付加する。
    """
    df = df.copy()
    df[key_col] = df[key_col].map(lambda x: name_map.get(x, x))

    has_content = CONTENT_COL in df.columns

    def most_frequent(s):
        vc = s[s != ""].value_counts()
        return vc.index[0] if len(vc) > 0 else ""

    if group_by_content and has_content:
        _merge_only = merge_only_payees or set()
        mask_merge = df[key_col].isin(_merge_only)
        parts = []

        # merge_only_payees → key_col のみで集計（取引内容を問わずまとめる）
        if mask_merge.any():
            df_m = df[mask_merge]
            amt = df_m.groupby(key_col, as_index=False)[amount_col].sum()
            cnt = df_m.groupby(key_col, as_index=False)[CONTENT_COL].agg(most_frequent)
            parts.append(amt.merge(cnt, on=key_col)[[key_col, CONTENT_COL, amount_col]])

        # それ以外 → (key_col, CONTENT_COL) の組み合わせで集計
        if (~mask_merge).any():
            df_s = df[~mask_merge]
            agg = df_s.groupby([key_col, CONTENT_COL], as_index=False)[amount_col].sum()
            parts.append(agg[[key_col, CONTENT_COL, amount_col]])

        result = (
            pd.concat(parts, ignore_index=True)
            .sort_values(amount_col, ascending=False)
            .reset_index(drop=True)
        ) if parts else pd.DataFrame(columns=[key_col, CONTENT_COL, amount_col])

    else:
        # 従来の動作: key_col のみで集計
        amount_result = (
            df.groupby(key_col, as_index=False)[amount_col]
            .sum()
            .sort_values(amount_col, ascending=False)
            .reset_index(drop=True)
        )
        if has_content:
            content_result = df.groupby(key_col, as_index=False)[CONTENT_COL].agg(most_frequent)
            result = amount_result.merge(content_result, on=key_col)
            result = result[[key_col, CONTENT_COL, amount_col]]
        else:
            result = amount_result

    result.index = range(1, len(result) + 1)
    return result


def to_excel_bytes(df: pd.DataFrame, title: str, amount_col: str) -> bytes:
    """集計済み DataFrame を整形した Excel バイト列に変換する。列構成は df.columns に従う。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "内訳"

    # ── スタイル定義 ──
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font  = Font(bold=True, size=13)
    total_font  = Font(bold=True, size=11)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")
    right  = Alignment(horizontal="right",  vertical="center")

    cols = list(df.columns)

    # ── タイトル行 ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = title_font
    title_cell.alignment = center

    # ── ヘッダー行 ──
    for c_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=2, column=c_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # ── データ行 ──
    for r_idx, row in enumerate(df.itertuples(index=False), start=3):
        for c_idx, (col_name, value) in enumerate(zip(cols, row), start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border
            if col_name == amount_col:
                cell.number_format = "#,##0"
                cell.alignment = right
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=True)

    # ── 合計行 ──
    total_row = len(df) + 3
    amount_col_idx = cols.index(amount_col) + 1  # 1始まり
    for c_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=total_row, column=c_idx)
        cell.border = border
        cell.font = total_font
        if c_idx == 1:
            cell.value = "合　計"
            cell.alignment = center
        elif col_name == amount_col:
            cell.value = df[amount_col].sum()
            cell.number_format = "#,##0"
            cell.alignment = right

    # ── 列幅調整 ──
    col_widths = {0: 32, 1: 22, 2: 16}  # 先頭から順に幅設定（3列想定、余剰は14）
    for i in range(len(cols)):
        ws.column_dimensions[get_column_letter(i + 1)].width = col_widths.get(i, 14)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────
# Streamlit アプリ本体
# ─────────────────────────────────────────────
def main():
    st.set_page_config(page_title="勘定科目内訳書作成ツール", page_icon="📋", layout="wide")
    st.title("📋 勘定科目内訳書作成ツール")

    # ── モード選択 ──
    mode_label = st.radio(
        "対象科目を選択してください",
        list(MODES.keys()),
        horizontal=True,
    )
    cfg = MODES[mode_label]
    key_col                  = cfg["key_label"]
    amount_col               = cfg["amount_label"]
    title                    = cfg["title"]
    amount_col_idx           = cfg["amount_col_idx"]
    opposite_col_idx         = cfg["opposite_col_idx"]
    csv_amount_col_name      = cfg["csv_amount_col_name"]
    csv_opposite_col_name    = cfg["csv_opposite_col_name"]
    default_top_n            = cfg["default_top_n"]
    taxable_kaku_codes       = cfg["taxable_kaku_codes"]
    group_map                = cfg["group_map"]
    group_by_content         = cfg.get("group_by_content", False)
    ledger_tolerance         = cfg["ledger_tolerance"]

    st.divider()

    # ── ファイルアップロード ──
    st.subheader("① ファイルのアップロード")

    col_upload, col_skip = st.columns([4, 1])
    with col_upload:
        uploaded = st.file_uploader(
            "JDL元帳ファイル（.xlsx / .xls / .csv）をアップロードしてください",
            type=["xlsx", "xls", "csv"],
            help=(
                "Excel: 列順 = 年月日 | 相手科目 | 摘要 | 借方 | 貸方 | 残高\n"
                "CSV(Shift-JIS): 1列目が「 #」の行のみ処理、7列目=摘要、14列目=借方、15列目=貸方"
            ),
        )
    with col_skip:
        skiprows = st.number_input(
            "スキップ行数（Excelのみ）",
            min_value=0,
            max_value=20,
            value=0,
            step=1,
            help="Excelファイルの先頭にある不要行の数を指定してください（CSVは無視されます）",
        )

    if uploaded is None:
        st.info(
            "JDL元帳ファイル（Excel または CSV）をアップロードすると処理を開始します。\n\n"
            "**Excel の列構成（位置固定）**\n"
            "1列目: 年月日 ／ 2列目: 相手科目 ／ 3列目: 摘要 ／ "
            "4列目: 借方 ／ 5列目: 貸方 ／ 6列目: 残高\n\n"
            "**CSV の仕様（Shift-JIS）**\n"
            "1列目が「 #」で始まる行のみ処理 ／ 摘要・借方金額・貸方金額の列はヘッダー行から自動検出\n\n"
            f"**{mode_label}** は {'借方' if amount_col_idx == COL_DEBIT else '貸方'} を集計します。"
        )
        st.stop()

    # ── データ読み込み ──
    is_csv = uploaded.name.lower().endswith(".csv")
    ledger_total: float | None = None
    _taxinc_net_total: float = 0.0
    apply_tax_exclusion: bool = False
    try:
        if is_csv:
            raw_df, raw_all, _final_balance, _total_row, _taxinc_net_total, apply_tax_exclusion = load_csv_file(
                uploaded, csv_amount_col_name, csv_opposite_col_name,
                key_col, amount_col, taxable_kaku_codes, group_map
            )
            ledger_total = _final_balance  # 残高列の最終値を正解値として使用
        else:
            raw_df, raw_all, apply_tax_exclusion, _final_balance = load_jdl_excel(
                uploaded, amount_col_idx, opposite_col_idx,
                key_col, amount_col, taxable_kaku_codes, group_map,
                skiprows=int(skiprows)
            )
            ledger_total = _final_balance  # 残高列の最終値を正解値として使用
    except Exception as e:
        st.error(f"ファイルの読み込みに失敗しました: {e}")
        st.stop()

    if raw_df.empty:
        st.warning(
            "有効なデータが0件です。\n"
            + ("・1列目が「 #」で始まる行が存在するか確認してください\n" if is_csv else
               "・スキップ行数の設定を確認してください\n")
            + f"・{'借方' if (csv_amount_col_name if is_csv else ('借方' if amount_col_idx == COL_DEBIT else '貸方')) in ('借方金額', '借方') else '貸方'}に金額が入力されている行があるか確認してください"
        )
        with st.expander("読み込んだ全データ（生）を確認する"):
            st.dataframe(raw_all, use_container_width=True)
        st.stop()

    # 税抜き / 税込み判定結果の表示
    if apply_tax_exclusion:
        st.info("💹 **税抜き法人**として集計します（消費税額振替の仕訳を検出 → 税抜き計算を適用）")
    else:
        st.info("💹 **税込み法人**として集計します（消費税額振替の仕訳なし → 金額をそのまま使用）")

    with st.expander("読み込んだ生データを確認する（抽出後）", expanded=False):
        sep_desc = "全角スペース（\u3000）" if is_csv else "スペース2つ以上"
        st.caption(f"摘要列を{sep_desc}で分割し、先頭部分を「{key_col}」として抽出しました")
        st.dataframe(raw_df, use_container_width=True)

    st.divider()

    # ── 自動統合（類似度0.75以上・件数多数決） ──
    st.subheader("② 自動名称統合（類似度 0.75 以上）")

    name_counts = raw_df[key_col].value_counts().to_dict()
    auto_name_map = auto_merge_by_frequency(
        list(raw_df[key_col].unique()), name_counts, threshold=0.75
    )

    # 自動統合を適用した作業用 DataFrame（以降の手動補正・集計はこちらを使う）
    working_df = raw_df.copy()
    working_df[key_col] = working_df[key_col].map(lambda x: auto_name_map.get(x, x))

    if auto_name_map:
        st.info(f"{len(auto_name_map)} 件の表記を自動統合しました（件数の多い表記を正式名称として採用）。")
        auto_map_rows = [
            {
                "統合前の表記": old,
                "統合後（正式名称）": new,
                "統合前の件数": name_counts.get(old, 0),
                "正式名称の件数": name_counts.get(new, 0),
            }
            for old, new in auto_name_map.items()
        ]
        with st.expander("自動統合の詳細を確認する", expanded=True):
            st.dataframe(pd.DataFrame(auto_map_rows), use_container_width=True, hide_index=True)
    else:
        st.success("自動統合の対象となる類似名称はありませんでした。")

    st.divider()

    # ── 手動 表記ゆれ補正 ──
    st.subheader("③ 手動 表記ゆれ補正（任意）")

    threshold = st.slider(
        "類似度のしきい値（低いほど広くマッチ）",
        min_value=0.50,
        max_value=1.00,
        value=SIMILARITY_THRESHOLD,
        step=0.01,
        format="%.2f",
    )

    # 自動統合後の名称リストで表記ゆれを検索
    unique_names = sorted(working_df[key_col].unique().tolist())
    similar_groups = find_similar_groups(unique_names, threshold)

    # セッションステートで名称マッピングを管理
    if "name_map" not in st.session_state:
        st.session_state.name_map = {}

    if not similar_groups:
        st.success("追加の表記ゆれは検出されませんでした。")
    else:
        st.warning(f"{len(similar_groups)} 件の表記ゆれグループを検出しました。統一後の名称を設定してください。")

        for base_name, group in similar_groups.items():
            with st.expander(f"グループ: **{base_name}** 他 {len(group) - 1} 件", expanded=True):
                st.write("検出されたバリアント:", group)
                unified = st.text_input(
                    "統一後の名称",
                    value=base_name,
                    key=f"unified_{base_name}",
                )
                apply = st.checkbox("このグループを統合する", value=True, key=f"apply_{base_name}")
                if apply:
                    for name in group:
                        st.session_state.name_map[name] = unified

    # 手動マッピング追加
    with st.expander("手動で名称を統一する（任意）", expanded=False):
        st.write("個別に名称を変更できます。")
        col1, col2, col3 = st.columns([3, 3, 1])
        with col1:
            src_name = st.selectbox("変更前の名称", [""] + unique_names, key="manual_src")
        with col2:
            dst_name = st.text_input("変更後の名称", key="manual_dst")
        with col3:
            st.write("")
            st.write("")
            if st.button("追加", use_container_width=True):
                if src_name and dst_name:
                    st.session_state.name_map[src_name] = dst_name
                    st.success(f"「{src_name}」→「{dst_name}」を登録しました。")

        if st.session_state.name_map:
            st.write("現在の名称マッピング:")
            map_df = pd.DataFrame(
                list(st.session_state.name_map.items()),
                columns=["変更前", "変更後"],
            )
            st.dataframe(map_df, use_container_width=True, hide_index=True)

            if st.button("マッピングをリセット"):
                st.session_state.name_map = {}
                st.rerun()

    st.divider()

    # ── 名寄せ設定 ──
    st.subheader("④ 名寄せ設定（任意）")

    # GROUP_MAP の統合対象（values）と「寮費」は取引内容を問わず payee 単位でまとめる
    merge_only_payees = (set(group_map.values()) | {"寮費"}) if group_by_content else None

    # 集計後の支払先一覧を取得（ドロップダウンの選択肢 & 以降のプレビューに共用）
    result_df = aggregate(
        working_df, key_col, amount_col, st.session_state.name_map,
        group_by_content=group_by_content,
        merge_only_payees=merge_only_payees,
    )
    total_all = result_df[amount_col].sum()

    aggregated_names = result_df[key_col].unique().tolist()

    st.caption("集計後の支払先を2つ選んで統合できます。件数が多い方の名称が正式名称として採用されます。")
    col_a, col_b, col_btn = st.columns([3, 3, 1])
    with col_a:
        merge_name_a = st.selectbox(
            "統合元 A", [""] + aggregated_names, key="merge_a",
        )
    with col_b:
        merge_name_b = st.selectbox(
            "統合元 B", [""] + aggregated_names, key="merge_b",
        )
    with col_btn:
        st.write("")
        st.write("")
        if st.button("統合する", use_container_width=True, key="btn_merge"):
            if merge_name_a and merge_name_b and merge_name_a != merge_name_b:
                # working_df に現在の name_map を適用した状態で件数を比較
                applied_names = working_df[key_col].map(
                    lambda x: st.session_state.name_map.get(x, x)
                )
                count_a = (applied_names == merge_name_a).sum()
                count_b = (applied_names == merge_name_b).sum()
                canonical = merge_name_a if count_a >= count_b else merge_name_b
                other     = merge_name_b if count_a >= count_b else merge_name_a
                st.session_state.name_map[other] = canonical
                st.rerun()
            elif merge_name_a == merge_name_b and merge_name_a:
                st.warning("同じ名称が選択されています。別の組み合わせを選んでください。")

    if st.session_state.name_map:
        with st.expander("現在の名寄せ設定を確認する", expanded=False):
            map_df = pd.DataFrame(
                list(st.session_state.name_map.items()),
                columns=["統合前", "統合後（正式名称）"],
            )
            st.dataframe(map_df, use_container_width=True, hide_index=True)
            if st.button("名寄せをリセット", key="btn_reset_merge"):
                st.session_state.name_map = {}
                st.rerun()

    st.divider()

    # ── 集計結果プレビュー ──
    st.subheader("⑤ 集計結果プレビュー")

    col_topn, _ = st.columns([1, 3])
    with col_topn:
        top_n = st.number_input(
            "上位N件で集計（0 = 全件表示）",
            min_value=0,
            max_value=500,
            value=default_top_n,
            step=1,
            help="指定件数を超える分は「その他（N件）」1行にまとめます",
        )

    # ── 残高チェック（CSV・Excel共通） ──
    if ledger_total is not None:
        # 税込み・税抜きに関わらず集計合計 vs 残高列の最終値で照合
        check_total = total_all
        diff = check_total - ledger_total
        if abs(diff) <= ledger_tolerance:
            st.success(f"✅ 元帳とほぼ一致しています（誤差：{diff:+,.0f}円）")
        else:
            st.error(f"⚠️ 差額 {diff:+,.0f}円 が解消できませんでした")

    total_count = len(result_df)
    result_df = collapse_to_top_n(result_df, key_col, amount_col, int(top_n))

    # 表示列順：収入先名 | 取引内容（あれば） | 金額
    display_cols = [c for c in [key_col, CONTENT_COL, amount_col] if c in result_df.columns]

    col_left, col_right = st.columns([3, 1])
    with col_left:
        st.dataframe(
            result_df[display_cols].style.format({amount_col: "{:,.0f}"}),
            use_container_width=True,
        )
    with col_right:
        st.metric("合計金額", f"¥{total_all:,.0f}")
        st.metric("件数（集計前）", f"{total_count} 件")
        if int(top_n) > 0 and total_count > int(top_n):
            st.metric("表示件数", f"上位{int(top_n)}件＋その他")

    st.divider()

    # ── Excelダウンロード ──
    st.subheader("⑥ Excelダウンロード")

    excel_bytes = to_excel_bytes(result_df[display_cols], title, amount_col)

    st.download_button(
        label="📥 Excelをダウンロード",
        data=excel_bytes,
        file_name=f"{mode_label}_内訳書.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        type="primary",
    )


if __name__ == "__main__":
    main()
